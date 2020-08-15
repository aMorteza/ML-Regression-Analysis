import sys
import scipy.io as sio
from pprint import pprint
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from pathlib import Path

xlsx_file = Path('ENB2012_data.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 
sheet = wb_obj.active

training_data = []
for row in sheet.iter_rows(min_row=2, max_row=600, min_col=1, max_col=10):
    data = []
    for cell in row:
        data.append(cell.value)
    training_data.append(data)

testing_data = []
for row in sheet.iter_rows(min_row=600, max_row=769, min_col=1, max_col=10):
    data = []
    for cell in row:
        data.append(cell.value)
    testing_data.append(data)