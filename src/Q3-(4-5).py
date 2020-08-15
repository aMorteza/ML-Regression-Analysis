import sys
import scipy.io as sio
from pprint import pprint
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from pathlib import Path

from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.multioutput import MultiOutputRegressor
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.model_selection import cross_val_score
from sklearn.metrics import accuracy_score
from sklearn.metrics import r2_score



xlsx_file = Path('ENB2012_data.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 
sheet = wb_obj.active

X = []
for row in sheet.iter_rows(min_row=2, max_row=769, min_col=1, max_col=8):
    data = []
    for cell in row:
        data.append(cell.value)
    X.append(data)
X = np.array(X)
y = []
for row in sheet.iter_rows(min_row=2, max_row=769, min_col=9, max_col=10):
    data = []
    for cell in row:
        data.append(cell.value)
    y.append(data)
y = np.array(y)

X_train, X_test, y_train, y_test = train_test_split(
    X, y, train_size=600, test_size=168, shuffle=False)


def cv_10folds(X):
	n = X.shape[0]
	i = 1
	while i <= 10:
		idx = np.arange(n * (i - 1) / 10, n * i / 10, dtype=int)
		yield idx, idx
		i += 1


multi_output_regr = MultiOutputRegressor(GradientBoostingRegressor(alpha=0.007, validation_fraction=0.1,  n_iter_no_change=45))

cv = cv_10folds(X)
cross_val_score = cross_val_score(multi_output_regr, X, y, cv=cv)
print("cross_val_score:", cross_val_score)

multi_output_regr.fit(X_train, y_train)

# Predict on new data
y_multi_gd = multi_output_regr.predict(X_test)

plt.figure()
plt.scatter(y_test[:, 0], y_test[:, 1], edgecolor='k', c="r", label="Data")
plt.scatter(y_multi_gd[:, 0], y_multi_gd[:, 1], edgecolor='w', c="b", marker="s", label="Multi GD score=%.2f" % multi_output_regr.score(X_test, y_test))
plt.xlabel("target 1")
plt.ylabel("target 2")
plt.title("MultiOutputRegressor")
plt.legend()
plt.show()